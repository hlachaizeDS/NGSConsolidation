#import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.formatting.rule import DataBarRule
import os
import xlrd
from small_functions import *

directory = 'To_Consolidate'

stats_L = []
norm_L = []
for filename in os.listdir(os.path.join(os.getcwd(), directory)):
    if filename[-5:]=='.xlsm':
        print(filename)
        workbook=xlrd.open_workbook(directory + "\\" + filename)
        stats_L += get_values_from_sheet_xlrd(workbook,'Preprocessing stats')[1]
        norm_L += get_values_from_sheet_xlrd(workbook,'Per position norm')[1]
print(stats_L)
wb = load_workbook(filename='pivot_table.xlsm', read_only=False, keep_vba=True)

ws = wb.create_sheet(title='Preprocessing stats')
#stats_df = pd.concat(stats_L)

for r in stats_L:
    ws.append(r)

for col in ws.iter_cols():
    print(col)
    if '%' in col[0].value or 'Purity' in col[0].value:
        print('%')
        for cell in col:
            ws[cell.coordinate].number_format = '0%'
    elif 'FileID' not in col[0].value:
        for cell in col:
            ws[cell.coordinate].number_format = '# ##0'

            ws = wb.create_sheet(title='Per position norm')
            #display(norm_L[0])
            #norm_df = pd.concat(norm_L)
            print('before appending')
            for r in norm_L:

                ws.append(r)
            print('after appending')
            for col in ws.iter_cols():
                if col[0].value not in ['NGS', 'FileID', 'Position', '>=1000Reads', 'Cycle', 'IsCoreSequence', 'IsScar', 'SequenceLength', 'FullSequenceLength', 'Base', 'Library', 'AlignedReads', 'Depth', 'PolyN', 'PureReads','DistanceToSecStruct',
                                        'SSMinimumFreeEnergy', 'SSThermoEnsembleFreeEnergy', 'DimerSSMinimumFreeEnergy', 'DimerSSThermoEnsembleFreeEnergy',
                                        'Dimer1DistanceToSecStruct', 'Dimer2DistanceToSecStruct', 'DimerSSDeltaG', '>=3GPatternNumber', 'DistanceTo>=3G', 'Temperature (°C)',
                                        '[Enzyme] (uM)', '[Nucleotide] (uM)', 'Scale (pmol)', 'ElongationTime (s)', 'WellColumn', 'OP2Purity']:
                    for cell in col:
                        ws[cell.coordinate].number_format = '0.0%'
print('bef consolidation')
wb.save(directory + '_consolidation.xlsm')


def get_excel_sheet_values_xlrd(sheet):
    sheet_values=[]
    for row in range(0,sheet.nrows):
        column=[]
        for col in range(0,sheet.ncols):
            column.append(sheet.cell(row,col).value)
        sheet_values.append(column)

    return sheet_values

def get_values_from_sheet_xlrd(workbook,sheetname):
    if sheetname not in workbook.sheet_names():
        print("Tab " + sheetname + " not found")
        return[0,[]]
    else:
        tab = workbook.sheet_by_name(sheetname)
        values = get_excel_sheet_values_xlrd(tab)
        return [1,values]
