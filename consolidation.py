import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.formatting.rule import DataBarRule
import os
from tkinter import filedialog
from tkinter import *

root = Tk()
root.withdraw()
directory = filedialog.askdirectory(initialdir=os.path.dirname(os.path.realpath(__file__)),title="Select folder with NGS files to consolidate")


stats_L = []
norm_L = []
for filename in os.listdir(os.path.join(os.getcwd(), directory)):
    if filename[-5:]=='.xlsm':
        print(filename)
        file = pd.ExcelFile(os.path.join(directory, filename))
        stats_L += [pd.read_excel(file, sheet_name='Preprocessing stats')]
        norm_L += [pd.read_excel(file, sheet_name='Per position norm')]


print(stats_L)

wb = load_workbook(filename='pivot_table.xlsm', read_only=False, keep_vba=True)

print('Copying Preprocessing stats')
ws = wb.create_sheet(title='Preprocessing stats')
stats_df = pd.concat(stats_L)


rows_total=len(stats_df.index)
counter=0
evolution_percent=10
for r in dataframe_to_rows(stats_df, header=True, index=False):
    ws.append(r)
    counter+=1
    percent=100*(counter/rows_total)
    if percent>evolution_percent:
        print(str(evolution_percent) + "%")
        evolution_percent+=10

print("Changing cells format")
for col in ws.iter_cols():
    if '%' in col[0].value or 'Purity' in col[0].value:
        for cell in col:
            ws[cell.coordinate].number_format = '0%'
    elif 'FileID' not in col[0].value:
        for cell in col:
            ws[cell.coordinate].number_format = '# ##0'

ws = wb.create_sheet(title='Per position norm')
#display(norm_L[0])
norm_df = pd.concat(norm_L)
print('-------------------------')
print('Copying Per position norm')
rows_total=len(norm_df.index)
counter=0
evolution_percent=10
for r in dataframe_to_rows(norm_df, header=True, index=False):
    ws.append(r)
    counter+=1
    percent=100*(counter/rows_total)
    if percent>evolution_percent:
        print(str(evolution_percent) + "%")
        evolution_percent+=10
print("Changing cells format")
for col in ws.iter_cols():
    if col[0].value not in ['NGS', 'FileID', 'Position', '>=1000Reads', 'Cycle', 'IsCoreSequence', 'IsScar', 'SequenceLength', 'FullSequenceLength', 'Base', 'Library', 'AlignedReads', 'Depth', 'PolyN', 'PureReads','DistanceToSecStruct',
                                        'SSMinimumFreeEnergy', 'SSThermoEnsembleFreeEnergy', 'DimerSSMinimumFreeEnergy', 'DimerSSThermoEnsembleFreeEnergy',
                                        'Dimer1DistanceToSecStruct', 'Dimer2DistanceToSecStruct', 'DimerSSDeltaG', '>=3GPatternNumber', 'DistanceTo>=3G', 'Temperature (°C)',
                                        '[Enzyme] (uM)', '[Nucleotide] (uM)', 'Scale (pmol)', 'ElongationTime (s)', 'WellColumn', 'OP2Purity']:
        for cell in col:
            ws[cell.coordinate].number_format = '0.0%'
print('-------------------------')
print("Saving")
directory_name=os.path.basename(directory)
wb.save(directory + "\\" + directory_name + '_consolidation.xlsm')
